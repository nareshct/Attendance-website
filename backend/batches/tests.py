import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIRequestFactory, APITestCase, force_authenticate

from billing.models import BillingCycle, ClientInvoice, Payout
from courses.models import Course
from students.models import Student
from trainers.models import Trainer

from .models import Batch, BatchEnrollment, BatchInstallment, BatchPayout, BatchSession
from .services import all_batches_summary, batch_revenue_summary, create_batch_installments, source_report
from .views import (
    BatchEnrollmentViewSet,
    BatchInstallmentViewSet,
    BatchPayoutViewSet,
    BatchSessionViewSet,
    BatchViewSet,
    MyBatchesView,
)


def _make_batch(payment_type='one_time', fee=Decimal('4000'), total_classes=24, trainer_names=''):
    course = Course.objects.create(name='Scratch Programming', total_classes=24, rate_per_class=Decimal('1000'))
    return Batch.objects.create(
        name='Scratch — July batch', course=course, total_classes=total_classes,
        fee_per_student=fee, payment_type=payment_type, start_date=datetime.date(2026, 7, 1),
        trainer_names=trainer_names,
    )


def _make_student(name='Kid'):
    return Student.objects.create(name=name, grade='5', source_type='B2C')


def _make_trainer(username='trainer', name='Priya'):
    user = get_user_model().objects.create_user(username=username, password='x')
    return Trainer.objects.create(user=user, name=name, phone_number='0000000000', place='Here', default_rate_per_class=Decimal('100'))


class CreateBatchInstallmentsTests(TestCase):
    """Installments must always sum to exactly fee_per_student, and
    due_at_sessions milestones must be resolved from the batch's own
    total_classes, not tied to any per-student attendance."""

    def test_one_time_is_a_single_installment_due_immediately(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installments = create_batch_installments(enrollment)
        self.assertEqual(len(installments), 1)
        self.assertIsNone(installments[0].due_at_sessions)
        self.assertEqual(installments[0].amount, Decimal('4000.00'))

    def test_two_installments_even_split(self):
        batch = _make_batch(payment_type='two_installments', fee=Decimal('4000'), total_classes=24)
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installments = create_batch_installments(enrollment)
        self.assertEqual(len(installments), 2)
        self.assertIsNone(installments[0].due_at_sessions)
        self.assertEqual(installments[0].amount, Decimal('2000.00'))
        self.assertEqual(installments[1].due_at_sessions, 12)  # 50% of 24 sessions
        self.assertEqual(installments[1].amount, Decimal('2000.00'))
        self.assertEqual(sum(i.amount for i in installments), Decimal('4000.00'))

    def test_three_installments_uneven_split_absorbed_by_last(self):
        batch = _make_batch(payment_type='three_installments', fee=Decimal('1000'), total_classes=24)
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installments = create_batch_installments(enrollment)
        self.assertEqual(len(installments), 3)
        # 1000 / 3 = 333.33, 333.33, remainder absorbed by last -> 333.34
        self.assertEqual(installments[0].amount, Decimal('333.33'))
        self.assertEqual(installments[1].amount, Decimal('333.33'))
        self.assertEqual(installments[2].amount, Decimal('333.34'))
        self.assertEqual(sum(i.amount for i in installments), Decimal('1000.00'))
        self.assertEqual([i.due_at_sessions for i in installments], [None, 8, 16])  # 34%/67% of 24

    def test_four_installments_milestones(self):
        batch = _make_batch(payment_type='four_installments', fee=Decimal('4000'), total_classes=12)
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installments = create_batch_installments(enrollment)
        due_dates = [i.due_at_sessions for i in installments]
        self.assertEqual(due_dates, [None, 3, 6, 9])  # 25%/50%/75% of 12
        self.assertEqual(sum(i.amount for i in installments), Decimal('4000.00'))


class BatchRevenueSummaryTests(TestCase):
    """batch_revenue_summary() must reflect real enrollment/payment state, and
    must never touch BillingCycle/Payout/ClientInvoice — this is a separate
    revenue stream from the individual per-class billing system."""

    def test_revenue_reflects_enrolled_count_and_paid_status(self):
        batch = _make_batch(payment_type='two_installments', fee=Decimal('4000'))
        for i in range(3):
            enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(f'Kid {i}'), joined_date=datetime.date(2026, 7, 1))
            create_batch_installments(enrollment)

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['enrolled_count'], 3)
        self.assertEqual(summary['gross_expected'], Decimal('12000.00'))
        self.assertEqual(summary['collected'], Decimal('0.00'))
        self.assertEqual(summary['pending'], Decimal('12000.00'))

        # Mark one student's first installment paid.
        first_enrollment = BatchEnrollment.objects.filter(batch=batch).first()
        inst = first_enrollment.installments.get(sequence=1)
        inst.paid_status = 'paid'
        inst.save()

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['collected'], Decimal('2000.00'))
        self.assertEqual(summary['pending'], Decimal('10000.00'))

    def test_net_after_payout_sums_every_pending_or_paid_payout(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        BatchPayout.objects.create(batch=batch, recipient_name='Trainer 1', amount=Decimal('3000'))
        BatchPayout.objects.create(batch=batch, recipient_name='Trainer 2', amount=Decimal('2000'), paid_status='paid')
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        inst = create_batch_installments(enrollment)[0]
        inst.paid_status = 'paid'
        inst.save()

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['collected'], Decimal('4000.00'))
        self.assertEqual(summary['total_payouts'], Decimal('5000.00'), 'both payouts count, paid or not')
        self.assertEqual(summary['net_after_payout'], Decimal('-1000.00'))

    def test_net_after_payout_excludes_a_cancelled_payout(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        BatchPayout.objects.create(batch=batch, recipient_name='Trainer 1', amount=Decimal('3000'))
        BatchPayout.objects.create(batch=batch, recipient_name='Trainer 2', amount=Decimal('2000'), paid_status='cancelled')

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['total_payouts'], Decimal('3000.00'), 'a cancelled payout was never actually owed')

    def test_withdrawn_students_are_excluded(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        create_batch_installments(enrollment)
        enrollment.status = 'withdrawn'
        enrollment.save()

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['enrolled_count'], 0)
        self.assertEqual(summary['gross_expected'], Decimal('0.00'))

    def test_batches_never_create_billing_cycle_records(self):
        batch = _make_batch(payment_type='two_installments', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        create_batch_installments(enrollment)
        for inst in enrollment.installments.all():
            inst.paid_status = 'paid'
            inst.save()
        payout = BatchPayout.objects.create(batch=batch, recipient_name='Guest Instructor', amount=Decimal('1000'))
        payout.paid_status = 'paid'
        payout.save()

        self.assertEqual(BillingCycle.objects.count(), 0)
        self.assertEqual(Payout.objects.count(), 0)
        self.assertEqual(ClientInvoice.objects.count(), 0)


class BatchApiTests(APITestCase):
    """End-to-end through the real view layer: create a batch, enroll a
    student, mark payments/payout paid, log a session, withdraw."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='batch_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()
        self.course = Course.objects.create(name='Robotics', total_classes=24, rate_per_class=Decimal('1000'))

    def _req(self, method, path, data=None):
        request = getattr(self.factory, method)(path, data, format='json')
        force_authenticate(request, user=self.admin)
        return request

    def test_create_batch_enroll_student_and_pay(self):
        request = self._req('post', '/api/batches/', {
            'name': 'Robotics — August batch', 'course': self.course.id, 'total_classes': 24,
            'fee_per_student': '4000', 'payment_type': 'two_installments',
            'start_date': '2026-08-01', 'class_time': '17:00', 'class_days': 'MON,WED',
        })
        response = BatchViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        batch_id = response.data['id']

        student = _make_student('Enrolling Kid')
        request2 = self._req('post', '/api/batch-enrollments/', {'batch': batch_id, 'student': student.id})
        response2 = BatchEnrollmentViewSet.as_view({'post': 'create'})(request2)
        self.assertEqual(response2.status_code, 201, response2.data)
        self.assertEqual(len(response2.data['installments']), 2)

        # Duplicate enrollment must be rejected.
        request3 = self._req('post', '/api/batch-enrollments/', {'batch': batch_id, 'student': student.id})
        response3 = BatchEnrollmentViewSet.as_view({'post': 'create'})(request3)
        self.assertEqual(response3.status_code, 400)

        installment_id = response2.data['installments'][0]['id']
        request4 = self._req('post', f'/api/batch-installments/{installment_id}/mark_paid/', {})
        response4 = BatchInstallmentViewSet.as_view({'post': 'mark_paid'})(request4, pk=installment_id)
        self.assertEqual(response4.status_code, 200)
        self.assertEqual(response4.data['paid_status'], 'paid')

        request5 = self._req('get', f'/api/batches/{batch_id}/revenue/')
        response5 = BatchViewSet.as_view({'get': 'revenue'})(request5, pk=batch_id)
        self.assertEqual(response5.data['collected'], Decimal('2000.00'))

    def test_add_unregistered_name_creates_and_links_a_new_student(self):
        # A batch enrollment can be added by just a name — only the name is required,
        # everything else is optional lead capture — but it always resolves to a real,
        # registered Student now (see find_or_create_student), never a name-only row.
        batch = _make_batch(payment_type='one_time')
        request = self._req('post', '/api/batch-enrollments/', {
            'batch': batch.id, 'name': 'Walk-in Kid', 'phone_number': '9876543210',
            'lead_source': 'Instagram ad',
        })
        response = BatchEnrollmentViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        self.assertIsNotNone(response.data['student'])
        self.assertEqual(response.data['student_name'], 'Walk-in Kid')
        self.assertIsNotNone(response.data['student_id_code'])
        self.assertEqual(response.data['lead_source'], 'Instagram ad')
        self.assertEqual(len(response.data['installments']), 1)

        enrollment = BatchEnrollment.objects.get(id=response.data['id'])
        self.assertEqual(enrollment.display_name, 'Walk-in Kid')
        self.assertEqual(enrollment.student.source_type, 'B2C')
        self.assertEqual(enrollment.student.parent_phone_number, '9876543210')
        self.assertEqual(Student.objects.count(), 1)

    def test_add_unregistered_name_passes_through_parent_name_and_place(self):
        batch = _make_batch(payment_type='one_time')
        request = self._req('post', '/api/batch-enrollments/', {
            'batch': batch.id, 'name': 'Walk-in Kid', 'phone_number': '9876543210',
            'parent_name': 'Meera Rao', 'place': 'Chennai',
        })
        response = BatchEnrollmentViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        student = Student.objects.get(name='Walk-in Kid')
        self.assertEqual(student.parent_name, 'Meera Rao')
        self.assertEqual(student.place, 'Chennai')

    def test_add_unregistered_name_matches_existing_student_by_name_and_phone(self):
        # Someone with this exact name + phone is already registered — must link them,
        # not create a duplicate Student record.
        existing = Student.objects.create(name='Walk-in Kid', grade='6', parent_phone_number='98765 43210', source_type='B2C')
        batch = _make_batch(payment_type='one_time')
        request = self._req('post', '/api/batch-enrollments/', {
            'batch': batch.id, 'name': 'Walk-in Kid', 'phone_number': '9876543210',
        })
        response = BatchEnrollmentViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data['student'], existing.id)
        self.assertEqual(Student.objects.count(), 1)

    def test_enrollment_requires_either_student_or_a_name(self):
        batch = _make_batch(payment_type='one_time')
        request = self._req('post', '/api/batch-enrollments/', {'batch': batch.id})
        response = BatchEnrollmentViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 400)

    def test_cannot_provide_both_student_and_a_name(self):
        batch = _make_batch(payment_type='one_time')
        student = _make_student()
        request = self._req('post', '/api/batch-enrollments/', {
            'batch': batch.id, 'student': student.id, 'name': 'Someone Else',
        })
        response = BatchEnrollmentViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 400)

    def test_cannot_double_enroll_same_name_via_manual_add(self):
        # Manually re-submitting "Add student" for the same name+phone (e.g. a
        # double-click, or re-adding someone already added) resolves to the same
        # already-created Student both times (see find_or_create_student), so the
        # second submission is rejected as an existing-student duplicate.
        batch = _make_batch(payment_type='one_time')
        body = {'batch': batch.id, 'name': 'Dup Kid', 'phone_number': '90000 11111'}

        request1 = self._req('post', '/api/batch-enrollments/', body)
        response1 = BatchEnrollmentViewSet.as_view({'post': 'create'})(request1)
        self.assertEqual(response1.status_code, 201, response1.data)

        request2 = self._req('post', '/api/batch-enrollments/', body)
        response2 = BatchEnrollmentViewSet.as_view({'post': 'create'})(request2)
        self.assertEqual(response2.status_code, 400)
        self.assertEqual(BatchEnrollment.objects.filter(batch=batch, student__name='Dup Kid').count(), 1)

    def test_name_without_phone_can_be_added_more_than_once(self):
        # No phone means there's nothing reliable to dedup on — same policy as the import.
        batch = _make_batch(payment_type='one_time')
        body = {'batch': batch.id, 'name': 'No Phone Kid'}

        request1 = self._req('post', '/api/batch-enrollments/', body)
        response1 = BatchEnrollmentViewSet.as_view({'post': 'create'})(request1)
        self.assertEqual(response1.status_code, 201, response1.data)

        request2 = self._req('post', '/api/batch-enrollments/', body)
        response2 = BatchEnrollmentViewSet.as_view({'post': 'create'})(request2)
        self.assertEqual(response2.status_code, 201, response2.data)
        self.assertEqual(BatchEnrollment.objects.filter(batch=batch, student__name='No Phone Kid').count(), 2)

    def test_withdraw_cancels_pending_installments(self):
        batch = _make_batch(payment_type='two_installments')
        student = _make_student()
        enrollment = BatchEnrollment.objects.create(batch=batch, student=student, joined_date=datetime.date(2026, 7, 1))
        create_batch_installments(enrollment)

        request = self._req('post', f'/api/batch-enrollments/{enrollment.id}/withdraw/', {})
        response = BatchEnrollmentViewSet.as_view({'post': 'withdraw'})(request, pk=enrollment.id)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['status'], 'withdrawn')
        statuses = set(BatchInstallment.objects.filter(batch_enrollment=enrollment).values_list('paid_status', flat=True))
        self.assertEqual(statuses, {'cancelled'})

    def test_session_log(self):
        batch = _make_batch(payment_type='one_time')

        request = self._req('post', '/api/batch-sessions/', {
            'batch': batch.id, 'date': '2026-07-05', 'conducted_by_name': 'Priya (co-founder)',
            'topic_covered': 'Intro to blocks', 'recording_link': 'https://drive.google.com/xyz',
        })
        response = BatchSessionViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(BatchSession.objects.filter(batch=batch).count(), 1)

    def test_batch_can_have_a_separate_payout_per_person(self):
        # The scenario this replaced: a batch co-taught by two trainers needs an
        # independent payout entry for each, not one shared lump sum.
        batch = _make_batch(payment_type='one_time')

        request1 = self._req('post', '/api/batch-payouts/', {
            'batch': batch.id, 'recipient_name': 'Trainer 1', 'amount': '3000',
        })
        response1 = BatchPayoutViewSet.as_view({'post': 'create'})(request1)
        self.assertEqual(response1.status_code, 201, response1.data)

        request2 = self._req('post', '/api/batch-payouts/', {
            'batch': batch.id, 'recipient_name': 'Trainer 2', 'amount': '2000',
        })
        response2 = BatchPayoutViewSet.as_view({'post': 'create'})(request2)
        self.assertEqual(response2.status_code, 201, response2.data)

        payout1_id = response1.data['id']
        mark_request = self._req('post', f'/api/batch-payouts/{payout1_id}/mark_paid/', {})
        mark_response = BatchPayoutViewSet.as_view({'post': 'mark_paid'})(mark_request, pk=payout1_id)
        self.assertEqual(mark_response.status_code, 200)
        self.assertEqual(mark_response.data['paid_status'], 'paid')

        # Trainer 2's payout is untouched by marking Trainer 1's paid.
        payout2_id = response2.data['id']
        get_request = self._req('get', f'/api/batch-payouts/?batch={batch.id}')
        list_response = BatchPayoutViewSet.as_view({'get': 'list'})(get_request)
        by_id = {p['id']: p for p in list_response.data['results']}
        self.assertEqual(by_id[payout1_id]['paid_status'], 'paid')
        self.assertEqual(by_id[payout2_id]['paid_status'], 'pending')

    def test_admin_can_cancel_a_pending_batch_payout(self):
        batch = _make_batch(payment_type='one_time')
        payout = BatchPayout.objects.create(batch=batch, recipient_name='Trainer 1', amount=Decimal('3000'))

        request = self._req('post', f'/api/batch-payouts/{payout.id}/cancel/', {})
        response = BatchPayoutViewSet.as_view({'post': 'cancel'})(request, pk=payout.id)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['paid_status'], 'cancelled')

    def test_cannot_cancel_an_already_paid_batch_payout(self):
        batch = _make_batch(payment_type='one_time')
        payout = BatchPayout.objects.create(batch=batch, recipient_name='Trainer 1', amount=Decimal('3000'), paid_status='paid')

        request = self._req('post', f'/api/batch-payouts/{payout.id}/cancel/', {})
        response = BatchPayoutViewSet.as_view({'post': 'cancel'})(request, pk=payout.id)
        self.assertEqual(response.status_code, 400)

    def test_batch_can_have_any_number_of_free_text_trainer_names(self):
        # No cap, and not restricted to the registered Trainer table — "any buddy" can
        # be listed, split schedules or larger co-teaching groups are both allowed.
        names = ['Priya (co-founder)', 'Ravi', 'Guest Instructor', 'Neha Suresh', 'Anjali', 'Karthik']
        request = self._req('post', '/api/batches/', {
            'name': 'Robotics — co-taught batch', 'course': self.course.id, 'total_classes': 24,
            'fee_per_student': '4000', 'payment_type': 'one_time', 'start_date': '2026-08-01',
            'class_time': '17:00', 'class_days': 'MON,WED',
            'trainer_names': ', '.join(names),
        })
        response = BatchViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        # validate_trainer_names strips stray whitespace around each comma-separated entry.
        self.assertEqual(response.data['trainer_names'], ','.join(names))


class BatchImportTests(APITestCase):
    """Bulk-enrolling students into a batch from an Excel export of a Google Form's
    responses (ad-driven batch sign-ups) — see services.import_students_from_excel."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='import_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()
        self.batch = _make_batch(payment_type='two_installments', fee=Decimal('4000'))

    def _make_excel(self, rows, headers=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers or ['Name', 'Phone Number', 'Occupation', 'Email', 'How did you know about us?', 'Payment Status'])
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return SimpleUploadedFile(
            'import.xlsx', buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _import(self, batch_id, file_obj):
        request = self.factory.post(f'/api/batches/{batch_id}/import_students/', {'file': file_obj}, format='multipart')
        force_authenticate(request, user=self.admin)
        return BatchViewSet.as_view({'post': 'import_students'})(request, pk=batch_id)

    def test_import_creates_and_links_registered_students(self):
        file_obj = self._make_excel([
            ['Asha Kumar', '9876543210', 'Engineer', 'asha@example.com', 'Instagram ad', 'Paid'],
            ['Vikram Rao', '9123456789', '', '', 'Referral', ''],
        ])
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['enrolled'], 2)
        self.assertEqual(response.data['marked_paid'], 1)
        self.assertEqual(response.data['skipped'], [])

        self.assertEqual(Student.objects.count(), 2)

        enrollment = BatchEnrollment.objects.get(batch=self.batch, student__name='Asha Kumar')
        self.assertIsNotNone(enrollment.student)
        self.assertEqual(enrollment.student.name, 'Asha Kumar')
        self.assertEqual(enrollment.student.source_type, 'B2C')
        self.assertEqual(enrollment.display_name, 'Asha Kumar')
        self.assertEqual(enrollment.occupation, 'Engineer')
        self.assertEqual(enrollment.lead_source, 'Instagram ad')
        installments = list(enrollment.installments.order_by('sequence'))
        self.assertEqual(installments[0].paid_status, 'paid')
        self.assertEqual(installments[1].paid_status, 'pending')

        vikram_enrollment = BatchEnrollment.objects.get(batch=self.batch, student__name='Vikram Rao')
        self.assertIsNotNone(vikram_enrollment.student)
        self.assertEqual(
            list(vikram_enrollment.installments.order_by('sequence').values_list('paid_status', flat=True)),
            ['pending', 'pending'],
        )

    def test_import_links_an_existing_registered_student_by_name_and_phone(self):
        # Someone with this exact name + phone is already a registered student —
        # import must link them, not create a duplicate Student record.
        existing = Student.objects.create(name='Asha Kumar', grade='6', parent_phone_number='98765 43210', source_type='B2C')
        file_obj = self._make_excel([['Asha Kumar', '9876543210', '', '', '', '']])
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['enrolled'], 1)
        self.assertEqual(Student.objects.count(), 1)
        enrollment = BatchEnrollment.objects.get(batch=self.batch)
        self.assertEqual(enrollment.student_id, existing.id)
        self.assertEqual(enrollment.student.name, 'Asha Kumar')

    def test_import_grade_column_is_used_when_creating_a_new_student(self):
        file_obj = self._make_excel(
            [['New Kid', '9876543210', '7', '', '', '']],
            headers=['Name', 'Phone Number', 'Grade', 'Occupation', 'Email', 'How did you know about us?', 'Payment Status'],
        )
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        student = Student.objects.get(name='New Kid')
        self.assertEqual(student.grade, '7')

    def test_import_parent_name_and_place_columns_are_used_when_creating_a_new_student(self):
        file_obj = self._make_excel(
            [['New Kid', '9876543210', 'Meera Rao', 'Chennai']],
            headers=['Name', 'Phone Number', 'Parent Name', 'Place'],
        )
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        student = Student.objects.get(name='New Kid')
        self.assertEqual(student.parent_name, 'Meera Rao')
        self.assertEqual(student.place, 'Chennai')

    def test_import_parent_name_and_place_are_not_used_when_matching_an_existing_student(self):
        # A match must never silently overwrite the existing student's own details with
        # whatever happened to be typed on this particular import row.
        existing = Student.objects.create(
            name='Asha Kumar', grade='6', parent_phone_number='98765 43210', source_type='B2C',
            parent_name='Original Parent', place='Original Place',
        )
        file_obj = self._make_excel(
            [['Asha Kumar', '9876543210', 'Different Parent', 'Different Place']],
            headers=['Name', 'Phone Number', 'Parent Name', 'Place'],
        )
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        existing.refresh_from_db()
        self.assertEqual(existing.parent_name, 'Original Parent')
        self.assertEqual(existing.place, 'Original Place')

    def test_import_skips_row_missing_name(self):
        file_obj = self._make_excel([['', '9876543210', '', '', '', '']])
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['enrolled'], 0)
        self.assertEqual(len(response.data['skipped']), 1)
        self.assertIn('Name is required', response.data['skipped'][0]['reason'])

    def test_import_skips_duplicate_row_within_batch(self):
        file_obj = self._make_excel([
            ['Repeat Kid', '9000000000', '', '', '', ''],
            ['Repeat Kid', '9000000000', '', '', '', ''],
        ])
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['enrolled'], 1)
        self.assertEqual(len(response.data['skipped']), 1)
        self.assertIn('already enrolled', response.data['skipped'][0]['reason'])

    def test_import_rejects_file_missing_name_column(self):
        file_obj = self._make_excel([['x']], headers=['Random Column'])
        response = self._import(self.batch.id, file_obj)
        self.assertEqual(response.status_code, 400)

    def test_import_template_download(self):
        request = self.factory.get('/api/batches/import_template/')
        force_authenticate(request, user=self.admin)
        response = BatchViewSet.as_view({'get': 'import_template'})(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


class BatchEnrollmentEditingTests(APITestCase):
    """PATCH/DELETE on an enrollment, plus undoing a withdraw — previously the only
    ways to touch an existing enrollment were create/withdraw."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='edit_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()
        self.batch = _make_batch(payment_type='one_time')

    def _req(self, method, path, data=None):
        request = getattr(self.factory, method)(path, data, format='json')
        force_authenticate(request, user=self.admin)
        return request

    def test_patch_updates_lead_info_fields(self):
        # Name/phone aren't editable here anymore — they live on the linked Student now.
        # Occupation/email/source are the lead-tracking fields still on the enrollment.
        enrollment = BatchEnrollment.objects.create(
            batch=self.batch, student=_make_student('Regular Kid'), occupation='Student',
            joined_date=datetime.date(2026, 7, 1),
        )
        request = self._req('patch', f'/api/batch-enrollments/{enrollment.id}/', {'occupation': 'Engineer'})
        response = BatchEnrollmentViewSet.as_view({'patch': 'partial_update'})(request, pk=enrollment.id)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(response.data['occupation'], 'Engineer')
        self.assertEqual(response.data['student_name'], 'Regular Kid')

    def test_patch_cannot_reassign_batch_or_student(self):
        other_batch = _make_batch(payment_type='one_time')
        enrollment = BatchEnrollment.objects.create(
            batch=self.batch, student=_make_student('Kid'), joined_date=datetime.date(2026, 7, 1),
        )
        request = self._req('patch', f'/api/batch-enrollments/{enrollment.id}/', {'batch': other_batch.id})
        response = BatchEnrollmentViewSet.as_view({'patch': 'partial_update'})(request, pk=enrollment.id)
        self.assertEqual(response.status_code, 400)

    def test_delete_removes_enrollment(self):
        enrollment = BatchEnrollment.objects.create(
            batch=self.batch, student=_make_student('Kid'), joined_date=datetime.date(2026, 7, 1),
        )
        create_batch_installments(enrollment)
        request = self.factory.delete(f'/api/batch-enrollments/{enrollment.id}/')
        force_authenticate(request, user=self.admin)
        response = BatchEnrollmentViewSet.as_view({'delete': 'destroy'})(request, pk=enrollment.id)
        self.assertEqual(response.status_code, 204)
        self.assertFalse(BatchEnrollment.objects.filter(id=enrollment.id).exists())

    def test_reactivate_undoes_withdraw(self):
        batch = _make_batch(payment_type='two_installments')
        enrollment = BatchEnrollment.objects.create(
            batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1),
        )
        create_batch_installments(enrollment)

        withdraw_req = self._req('post', f'/api/batch-enrollments/{enrollment.id}/withdraw/', {})
        BatchEnrollmentViewSet.as_view({'post': 'withdraw'})(withdraw_req, pk=enrollment.id)
        enrollment.refresh_from_db()
        self.assertEqual(enrollment.status, 'withdrawn')
        self.assertTrue(enrollment.installments.filter(paid_status='cancelled').exists())

        reactivate_req = self._req('post', f'/api/batch-enrollments/{enrollment.id}/reactivate/', {})
        response = BatchEnrollmentViewSet.as_view({'post': 'reactivate'})(reactivate_req, pk=enrollment.id)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['status'], 'active')
        statuses = set(enrollment.installments.values_list('paid_status', flat=True))
        self.assertEqual(statuses, {'pending'})


class BatchLockedFieldsTests(APITestCase):
    """fee_per_student/payment_type must lock once a batch has enrollments — existing
    installments are never recalculated (see services.create_batch_installments)."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='lock_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()

    def _patch(self, batch_id, data):
        request = self.factory.patch(f'/api/batches/{batch_id}/', data, format='json')
        force_authenticate(request, user=self.admin)
        return BatchViewSet.as_view({'patch': 'partial_update'})(request, pk=batch_id)

    def test_cannot_change_fee_once_enrolled(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        response = self._patch(batch.id, {'fee_per_student': '5000'})
        self.assertEqual(response.status_code, 400)
        batch.refresh_from_db()
        self.assertEqual(batch.fee_per_student, Decimal('4000.00'))

    def test_cannot_change_payment_type_once_enrolled(self):
        batch = _make_batch(payment_type='one_time')
        BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        response = self._patch(batch.id, {'payment_type': 'two_installments'})
        self.assertEqual(response.status_code, 400)

    def test_can_still_change_fee_before_any_enrollment(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        response = self._patch(batch.id, {'fee_per_student': '5000'})
        self.assertEqual(response.status_code, 200, response.data)

    def test_can_still_change_other_fields_once_enrolled(self):
        batch = _make_batch(payment_type='one_time')
        BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        response = self._patch(batch.id, {'status': 'completed'})
        self.assertEqual(response.status_code, 200, response.data)


class BatchRefundAndInstallmentEditTests(APITestCase):
    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='refund_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()

    def _req(self, method, path, data=None):
        request = getattr(self.factory, method)(path, data, format='json')
        force_authenticate(request, user=self.admin)
        return request

    def test_withdraw_with_refund_reduces_collected(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]
        installment.paid_status = 'paid'
        installment.save()

        request = self._req(
            'post', f'/api/batch-enrollments/{enrollment.id}/withdraw/',
            {'refund_amount': '4000', 'refund_note': 'Full refund, moved cities'},
        )
        response = BatchEnrollmentViewSet.as_view({'post': 'withdraw'})(request, pk=enrollment.id)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(Decimal(response.data['refunded_amount']), Decimal('4000.00'))

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['collected'], Decimal('0.00'))
        self.assertEqual(summary['total_refunded'], Decimal('4000.00'))

    def test_withdraw_without_refund_leaves_collected_untouched(self):
        # Money genuinely collected before withdrawal stays counted as revenue unless
        # explicitly refunded.
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]
        installment.paid_status = 'paid'
        installment.save()

        request = self._req('post', f'/api/batch-enrollments/{enrollment.id}/withdraw/', {})
        BatchEnrollmentViewSet.as_view({'post': 'withdraw'})(request, pk=enrollment.id)

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['collected'], Decimal('4000.00'))

    def test_patch_pending_installment_amount(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]

        request = self._req('patch', f'/api/batch-installments/{installment.id}/', {'amount': '3500'})
        response = BatchInstallmentViewSet.as_view({'patch': 'partial_update'})(request, pk=installment.id)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(Decimal(response.data['amount']), Decimal('3500.00'))

    def test_cannot_patch_amount_of_paid_installment(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]
        installment.paid_status = 'paid'
        installment.save()

        request = self._req('patch', f'/api/batch-installments/{installment.id}/', {'amount': '1000'})
        response = BatchInstallmentViewSet.as_view({'patch': 'partial_update'})(request, pk=installment.id)
        self.assertEqual(response.status_code, 400)

    def test_negative_installment_amount_is_rejected(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]

        request = self._req('patch', f'/api/batch-installments/{installment.id}/', {'amount': '-500'})
        response = BatchInstallmentViewSet.as_view({'patch': 'partial_update'})(request, pk=installment.id)
        self.assertEqual(response.status_code, 400)
        self.assertIn('amount', response.data)


class BatchFeeValidationTests(APITestCase):
    def test_negative_fee_per_student_is_rejected(self):
        from .serializers import BatchSerializer

        course = Course.objects.create(name='Fee Validation Course', total_classes=10)
        serializer = BatchSerializer(data={
            'name': 'Bad Fee Batch', 'course': course.id, 'total_classes': 10,
            'fee_per_student': '-100', 'payment_type': 'one_time', 'start_date': '2026-01-01',
        })
        self.assertFalse(serializer.is_valid())
        self.assertIn('fee_per_student', serializer.errors)


class BatchPayoutValidationTests(APITestCase):
    """BatchPayout.amount previously had zero validation anywhere (unlike
    BatchInstallment.amount) — see MinValueValidator on the model field."""

    def test_negative_payout_amount_is_rejected(self):
        from .serializers import BatchPayoutSerializer

        batch = _make_batch()
        serializer = BatchPayoutSerializer(data={'batch': batch.id, 'recipient_name': 'Priya', 'amount': '-500'})
        self.assertFalse(serializer.is_valid())
        self.assertIn('amount', serializer.errors)


class BatchInstallmentStatusGuardTests(APITestCase):
    """mark_paid/revoke previously had no state-transition guard (unlike
    BatchPayoutViewSet's equivalents) — a cancelled (withdrawn-student)
    installment could be resurrected as paid, inflating collected revenue."""

    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='installment_guard_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()

    def test_cannot_mark_a_cancelled_installment_paid(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]
        installment.paid_status = 'cancelled'
        installment.save(update_fields=['paid_status'])

        request = self.factory.post(f'/api/batch-installments/{installment.id}/mark_paid/')
        force_authenticate(request, user=self.admin)
        response = BatchInstallmentViewSet.as_view({'post': 'mark_paid'})(request, pk=installment.id)

        self.assertEqual(response.status_code, 400)
        installment.refresh_from_db()
        self.assertEqual(installment.paid_status, 'cancelled')

    def test_cannot_revoke_a_pending_installment(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]

        request = self.factory.post(f'/api/batch-installments/{installment.id}/revoke/')
        force_authenticate(request, user=self.admin)
        response = BatchInstallmentViewSet.as_view({'post': 'revoke'})(request, pk=installment.id)

        self.assertEqual(response.status_code, 400)

    def test_can_still_revoke_a_paid_installment(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installment = create_batch_installments(enrollment)[0]
        installment.paid_status = 'paid'
        installment.paid_date = datetime.date(2026, 7, 2)
        installment.save(update_fields=['paid_status', 'paid_date'])

        request = self.factory.post(f'/api/batch-installments/{installment.id}/revoke/')
        force_authenticate(request, user=self.admin)
        response = BatchInstallmentViewSet.as_view({'post': 'revoke'})(request, pk=installment.id)

        self.assertEqual(response.status_code, 200, response.data)
        installment.refresh_from_db()
        self.assertEqual(installment.paid_status, 'pending')


class BatchEnrollmentDuplicateRaceTests(APITestCase):
    """Two near-simultaneous "add student" submissions for the same name+phone both
    resolve (via find_or_create_student) to the same Student, so unique_together
    ('batch', 'student') — see BatchEnrollment.Meta — is the real backstop against a
    double-enroll race, not just the app-level dedup check in
    BatchEnrollmentViewSet.create().
    """

    def test_db_rejects_a_duplicate_enrollment_for_the_same_student(self):
        from django.db import IntegrityError

        batch = _make_batch()
        student = _make_student('Same Student')
        BatchEnrollment.objects.create(batch=batch, student=student, joined_date=datetime.date(2026, 7, 1))
        with self.assertRaises(IntegrityError):
            BatchEnrollment.objects.create(batch=batch, student=student, joined_date=datetime.date(2026, 7, 1))

    def test_create_view_returns_a_clean_400_not_a_500_for_a_duplicate(self):
        admin = get_user_model().objects.create_user(username='dup_add_admin', password='x', is_staff=True)
        batch = _make_batch()
        factory = APIRequestFactory()

        def _create():
            request = factory.post('/api/batch-enrollments/', {
                'batch': batch.id, 'name': 'Race Guest', 'phone_number': '9111111111',
            })
            force_authenticate(request, user=admin)
            return BatchEnrollmentViewSet.as_view({'post': 'create'})(request)

        first = _create()
        self.assertEqual(first.status_code, 201, first.data)
        second = _create()
        self.assertEqual(second.status_code, 400)
        self.assertEqual(BatchEnrollment.objects.filter(batch=batch, student__name='Race Guest').count(), 1)


class BatchGrossExpectedReflectsCorrectionsTests(TestCase):
    """gross_expected previously used enrolled_count x fee_per_student, which
    silently drifted from reality once a pending installment's amount was
    corrected — see batch_revenue_summary()."""

    def test_gross_expected_reflects_a_discounted_pending_installment(self):
        batch = _make_batch(payment_type='two_installments', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        installments = create_batch_installments(enrollment)
        installments[0].paid_status = 'paid'
        installments[0].save(update_fields=['paid_status'])
        installments[1].amount = Decimal('1000.00')  # discounted from 2000
        installments[1].save(update_fields=['amount'])

        summary = batch_revenue_summary(batch)
        self.assertEqual(summary['gross_expected'], Decimal('3000.00'))  # 2000 paid + 1000 discounted pending


class BatchExportEscapesFormulaLookingFieldsTests(TestCase):
    """Free text (student name, occupation/email/source) originates from a public
    Google Form or the enroll form directly — a value starting with =+-@ is written
    as a defused literal string, mirroring reports.views.csv_safe().
    """

    def test_formula_looking_student_name_is_escaped_in_the_export(self):
        from .services import build_export_workbook

        batch = _make_batch()
        student = Student.objects.create(name='=SUM(A1:A9)', grade='5', source_type='B2C')
        BatchEnrollment.objects.create(batch=batch, student=student, joined_date=datetime.date(2026, 7, 1))
        buffer = build_export_workbook(batch)
        workbook = load_workbook(buffer)
        sheet = workbook.active
        name_cell = sheet.cell(row=2, column=1).value
        self.assertEqual(name_cell, "'=SUM(A1:A9)")

    def test_formula_looking_source_is_escaped_in_the_export(self):
        from .services import build_export_workbook

        batch = _make_batch()
        student = _make_student('Normal Kid')
        BatchEnrollment.objects.create(
            batch=batch, student=student, lead_source='=SUM(A1:A9)', joined_date=datetime.date(2026, 7, 1),
        )
        buffer = build_export_workbook(batch)
        workbook = load_workbook(buffer)
        sheet = workbook.active
        source_cell = sheet.cell(row=2, column=6).value
        self.assertEqual(source_cell, "'=SUM(A1:A9)")


class BatchImportFileSizeCapTests(APITestCase):
    def test_oversized_import_file_is_rejected(self):
        from .views import BatchViewSet

        admin = get_user_model().objects.create_user(username='import_size_admin', password='x', is_staff=True)
        batch = _make_batch()
        factory = APIRequestFactory()
        oversized = SimpleUploadedFile(
            'huge.xlsx', b'x' * (BatchViewSet.MAX_IMPORT_FILE_SIZE + 1),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        request = factory.post(f'/api/batches/{batch.id}/import_students/', {'file': oversized}, format='multipart')
        force_authenticate(request, user=admin)
        response = BatchViewSet.as_view({'post': 'import_students'})(request, pk=batch.id)
        self.assertEqual(response.status_code, 400)


class AllBatchesSummaryPerformanceTests(TestCase):
    def test_query_count_does_not_scale_with_batch_count(self):
        for i in range(8):
            batch = _make_batch(payment_type='two_installments', fee=Decimal('4000'))
            enrollment = BatchEnrollment.objects.create(batch=batch, student=_make_student(f'Kid {i}'), joined_date=datetime.date(2026, 7, 1))
            create_batch_installments(enrollment)

        with CaptureQueriesContext(connection) as ctx:
            summary = all_batches_summary()
        self.assertEqual(summary['batch_count'], 8)
        # A handful of aggregate queries total, not ~4-5 per batch.
        self.assertLess(len(ctx), 10)


class BatchDeleteTests(APITestCase):
    def test_delete_batch_is_blocked(self):
        admin = get_user_model().objects.create_user(username='delete_admin', password='x', is_staff=True)
        factory = APIRequestFactory()
        batch = _make_batch(payment_type='one_time')
        request = factory.delete(f'/api/batches/{batch.id}/')
        force_authenticate(request, user=admin)
        response = BatchViewSet.as_view({'delete': 'destroy'})(request, pk=batch.id)
        self.assertEqual(response.status_code, 405)
        self.assertTrue(Batch.objects.filter(id=batch.id).exists())


class SourceReportTests(TestCase):
    def test_source_report_aggregates_enrollments_by_source(self):
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        e1 = BatchEnrollment.objects.create(batch=batch, student=_make_student('A'), lead_source='Instagram ad', joined_date=datetime.date(2026, 7, 1))
        create_batch_installments(e1)
        e2 = BatchEnrollment.objects.create(batch=batch, student=_make_student('B'), lead_source='Instagram ad', joined_date=datetime.date(2026, 7, 1))
        inst2 = create_batch_installments(e2)[0]
        inst2.paid_status = 'paid'
        inst2.save()
        e3 = BatchEnrollment.objects.create(batch=batch, student=_make_student('C'), lead_source='Referral', joined_date=datetime.date(2026, 7, 1))
        create_batch_installments(e3)
        # A registered-student add with no source at all is still counted — grouped
        # under "Not recorded", not dropped.
        e4 = BatchEnrollment.objects.create(batch=batch, student=_make_student(), joined_date=datetime.date(2026, 7, 1))
        create_batch_installments(e4)

        report = {row['source']: row for row in source_report()}
        self.assertEqual(report['Instagram ad']['enrolled_count'], 2)
        self.assertEqual(report['Instagram ad']['collected'], Decimal('4000.00'))
        self.assertEqual(report['Instagram ad']['pending'], Decimal('4000.00'))
        self.assertEqual(report['Referral']['enrolled_count'], 1)
        self.assertEqual(report['Not recorded']['enrolled_count'], 1)
        self.assertEqual(set(report), {'Instagram ad', 'Referral', 'Not recorded'})


class ExportStudentsTests(APITestCase):
    def test_export_returns_xlsx_with_roster(self):
        admin = get_user_model().objects.create_user(username='export_admin', password='x', is_staff=True)
        factory = APIRequestFactory()
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        enrollment = BatchEnrollment.objects.create(
            batch=batch, student=_make_student('Exportable Kid'), joined_date=datetime.date(2026, 7, 1),
        )
        create_batch_installments(enrollment)

        request = factory.get(f'/api/batches/{batch.id}/export_students/')
        force_authenticate(request, user=admin)
        response = BatchViewSet.as_view({'get': 'export_students'})(request, pk=batch.id)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        workbook = load_workbook(BytesIO(b''.join(response.streaming_content) if response.streaming else response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], 'Name')
        self.assertEqual(rows[1][0], 'Exportable Kid')

    def test_export_includes_lead_info_for_a_registered_enrollment(self):
        # Occupation/Email/Source are captured on every enrollment now — the export
        # must show them for a linked-student row, not just for a legacy guest row.
        # Name/phone come from the linked Student itself.
        admin = get_user_model().objects.create_user(username='export_lead_admin', password='x', is_staff=True)
        factory = APIRequestFactory()
        batch = _make_batch(payment_type='one_time', fee=Decimal('4000'))
        student = Student.objects.create(
            name='Registered Kid', grade='5', source_type='B2C', parent_phone_number='9000000002',
        )
        enrollment = BatchEnrollment.objects.create(
            batch=batch, student=student, occupation='Teacher', contact_email='kid@example.com',
            lead_source='Referral', joined_date=datetime.date(2026, 7, 1),
        )
        create_batch_installments(enrollment)

        request = factory.get(f'/api/batches/{batch.id}/export_students/')
        force_authenticate(request, user=admin)
        response = BatchViewSet.as_view({'get': 'export_students'})(request, pk=batch.id)
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(b''.join(response.streaming_content) if response.streaming else response.content))
        header = list(workbook.active.iter_rows(values_only=True))[0]
        row = dict(zip(header, list(workbook.active.iter_rows(values_only=True))[1]))
        self.assertEqual(row['Name'], 'Registered Kid')
        self.assertEqual(row['Phone Number'], '9000000002')
        self.assertEqual(row['Occupation'], 'Teacher')
        self.assertEqual(row['Email'], 'kid@example.com')
        self.assertEqual(row['How did you know about us?'], 'Referral')


class AcceptingEnrollmentsTests(APITestCase):
    def setUp(self):
        self.admin = get_user_model().objects.create_user(username='closed_admin', password='x', is_staff=True)
        self.factory = APIRequestFactory()

    def test_closed_batch_rejects_manual_add(self):
        batch = _make_batch(payment_type='one_time')
        batch.accepting_enrollments = False
        batch.save()
        request = self.factory.post('/api/batch-enrollments/', {'batch': batch.id, 'name': 'Late Kid'}, format='json')
        force_authenticate(request, user=self.admin)
        response = BatchEnrollmentViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 400)
        self.assertIn('closed', response.data['detail'])

    def test_closed_batch_rejects_import(self):
        from batches.services import import_students_from_excel
        batch = _make_batch(payment_type='one_time')
        batch.accepting_enrollments = False
        batch.save()

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Phone Number', 'Occupation', 'Email', 'How did you know about us?', 'Payment Status'])
        sheet.append(['Late Kid', '9000000000', '', '', '', ''])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        with self.assertRaises(ValueError):
            import_students_from_excel(batch, buffer)


class TrainerSelfServiceTests(APITestCase):
    def setUp(self):
        self.factory = APIRequestFactory()
        self.trainer = _make_trainer(username='priya_login', name='Priya')
        self.other_trainer = _make_trainer(username='ravi_login', name='Ravi')
        self.admin = get_user_model().objects.create_user(username='ss_admin', password='x', is_staff=True)
        self.my_batch = _make_batch(payment_type='one_time', trainer_names='Priya, Ravi')
        self.other_batch = _make_batch(payment_type='one_time', trainer_names='Ravi')

    def test_my_batches_only_lists_batches_the_trainer_is_on(self):
        request = self.factory.get('/api/my-batches/')
        force_authenticate(request, user=self.trainer.user)
        response = MyBatchesView.as_view()(request)
        self.assertEqual(response.status_code, 200)
        ids = {b['id'] for b in response.data['results']} if 'results' in response.data else {b['id'] for b in response.data}
        self.assertIn(self.my_batch.id, ids)
        self.assertNotIn(self.other_batch.id, ids)

    def test_my_batches_never_exposes_a_co_trainers_payout(self):
        # self.my_batch is co-taught by Priya (self.trainer) and Ravi (self.other_trainer).
        BatchPayout.objects.create(batch=self.my_batch, recipient_name='Priya', amount=Decimal('3000'))
        BatchPayout.objects.create(batch=self.my_batch, recipient_name='Ravi', amount=Decimal('5000'))

        request = self.factory.get('/api/my-batches/')
        force_authenticate(request, user=self.trainer.user)
        response = MyBatchesView.as_view()(request)
        self.assertEqual(response.status_code, 200)

        rows = response.data['results'] if 'results' in response.data else response.data
        batch_row = next(b for b in rows if b['id'] == self.my_batch.id)
        recipients = {p['recipient_name'] for p in batch_row['payouts']}
        amounts = {p['amount'] for p in batch_row['payouts']}
        self.assertEqual(recipients, {'Priya'})
        self.assertNotIn(Decimal('5000.00'), amounts)

    def test_trainer_can_log_session_for_their_own_batch(self):
        request = self.factory.post('/api/batch-sessions/', {
            'batch': self.my_batch.id, 'date': '2026-07-05', 'conducted_by_name': 'Priya',
        }, format='json')
        force_authenticate(request, user=self.trainer.user)
        response = BatchSessionViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)

    def test_trainer_cannot_log_session_for_a_batch_they_are_not_on(self):
        request = self.factory.post('/api/batch-sessions/', {
            'batch': self.other_batch.id, 'date': '2026-07-05', 'conducted_by_name': 'Priya',
        }, format='json')
        force_authenticate(request, user=self.trainer.user)
        response = BatchSessionViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 403)

    def test_co_trainer_cannot_edit_a_session_they_did_not_log(self):
        # self.my_batch is co-taught by Priya (self.trainer) and Ravi (self.other_trainer).
        session = BatchSession.objects.create(
            batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya', created_by=self.trainer.user,
        )
        request = self.factory.patch(
            f'/api/batch-sessions/{session.id}/', {'topic_covered': 'hijacked'}, format='json',
        )
        force_authenticate(request, user=self.other_trainer.user)
        response = BatchSessionViewSet.as_view({'patch': 'partial_update'})(request, pk=session.id)
        self.assertEqual(response.status_code, 403)

    def test_co_trainer_cannot_delete_a_session_they_did_not_log(self):
        session = BatchSession.objects.create(
            batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya', created_by=self.trainer.user,
        )
        request = self.factory.delete(f'/api/batch-sessions/{session.id}/')
        force_authenticate(request, user=self.other_trainer.user)
        response = BatchSessionViewSet.as_view({'delete': 'destroy'})(request, pk=session.id)
        self.assertEqual(response.status_code, 403)
        self.assertTrue(BatchSession.objects.filter(id=session.id).exists())

    def test_trainer_can_edit_their_own_logged_session(self):
        session = BatchSession.objects.create(
            batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya', created_by=self.trainer.user,
        )
        request = self.factory.patch(
            f'/api/batch-sessions/{session.id}/', {'topic_covered': 'updated'}, format='json',
        )
        force_authenticate(request, user=self.trainer.user)
        response = BatchSessionViewSet.as_view({'patch': 'partial_update'})(request, pk=session.id)
        self.assertEqual(response.status_code, 200, response.data)

    def test_admin_can_edit_any_trainers_session(self):
        session = BatchSession.objects.create(
            batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya', created_by=self.trainer.user,
        )
        request = self.factory.patch(
            f'/api/batch-sessions/{session.id}/', {'topic_covered': 'admin edit'}, format='json',
        )
        force_authenticate(request, user=self.admin)
        response = BatchSessionViewSet.as_view({'patch': 'partial_update'})(request, pk=session.id)
        self.assertEqual(response.status_code, 200, response.data)

    def test_legacy_session_with_no_created_by_stays_editable_by_any_co_trainer(self):
        # created_by is null for sessions logged before that field existed — those
        # must not become permanently locked out for every co-trainer on the batch.
        session = BatchSession.objects.create(batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya')
        request = self.factory.patch(
            f'/api/batch-sessions/{session.id}/', {'topic_covered': 'legacy edit'}, format='json',
        )
        force_authenticate(request, user=self.other_trainer.user)
        response = BatchSessionViewSet.as_view({'patch': 'partial_update'})(request, pk=session.id)
        self.assertEqual(response.status_code, 200, response.data)

    def test_created_by_is_set_automatically_on_create(self):
        request = self.factory.post('/api/batch-sessions/', {
            'batch': self.my_batch.id, 'date': '2026-07-06', 'conducted_by_name': 'Priya',
        }, format='json')
        force_authenticate(request, user=self.trainer.user)
        response = BatchSessionViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 201, response.data)
        session = BatchSession.objects.get(id=response.data['id'])
        self.assertEqual(session.created_by_id, self.trainer.user.id)

    def test_can_edit_is_true_for_the_owning_trainer_and_false_for_a_co_trainer(self):
        session = BatchSession.objects.create(
            batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya', created_by=self.trainer.user,
        )

        request = self.factory.get(f'/api/batch-sessions/?batch={self.my_batch.id}')
        force_authenticate(request, user=self.trainer.user)
        response = BatchSessionViewSet.as_view({'get': 'list'})(request)
        row = next(r for r in response.data['results'] if r['id'] == session.id)
        self.assertTrue(row['can_edit'])

        request = self.factory.get(f'/api/batch-sessions/?batch={self.my_batch.id}')
        force_authenticate(request, user=self.other_trainer.user)
        response = BatchSessionViewSet.as_view({'get': 'list'})(request)
        row = next(r for r in response.data['results'] if r['id'] == session.id)
        self.assertFalse(row['can_edit'])

    def test_can_edit_is_true_for_a_legacy_session_with_no_created_by(self):
        session = BatchSession.objects.create(batch=self.my_batch, date=datetime.date(2026, 7, 5), conducted_by_name='Priya')
        request = self.factory.get(f'/api/batch-sessions/?batch={self.my_batch.id}')
        force_authenticate(request, user=self.other_trainer.user)
        response = BatchSessionViewSet.as_view({'get': 'list'})(request)
        row = next(r for r in response.data['results'] if r['id'] == session.id)
        self.assertTrue(row['can_edit'])

    def test_trainer_only_sees_their_own_payouts(self):
        BatchPayout.objects.create(batch=self.my_batch, recipient_name='Priya', amount=Decimal('1000'))
        BatchPayout.objects.create(batch=self.my_batch, recipient_name='Ravi', amount=Decimal('2000'))

        request = self.factory.get(f'/api/batch-payouts/?batch={self.my_batch.id}')
        force_authenticate(request, user=self.trainer.user)
        response = BatchPayoutViewSet.as_view({'get': 'list'})(request)
        self.assertEqual(response.status_code, 200)
        names = {p['recipient_name'] for p in response.data['results']}
        self.assertEqual(names, {'Priya'})

    def test_trainer_cannot_create_a_payout(self):
        request = self.factory.post('/api/batch-payouts/', {
            'batch': self.my_batch.id, 'recipient_name': 'Priya', 'amount': '1000',
        }, format='json')
        force_authenticate(request, user=self.trainer.user)
        response = BatchPayoutViewSet.as_view({'post': 'create'})(request)
        self.assertEqual(response.status_code, 403)


class BatchEnrollmentReportCertificateTests(APITestCase):
    """report/certificate mirror EnrollmentViewSet.report/.certificate (see
    enrollments/tests.py) but scoped to a BatchEnrollment: any admin, or a
    trainer whose name is on the batch, can pull the report; the certificate
    additionally requires the whole Batch to be 'completed' and the specific
    BatchEnrollment to still be 'active'."""

    def setUp(self):
        self.factory = APIRequestFactory()
        self.admin = get_user_model().objects.create_user(username='cert_admin', password='x', is_staff=True)
        self.trainer = _make_trainer(username='cert_trainer', name='Priya')
        self.other_trainer = _make_trainer(username='cert_other_trainer', name='Ravi')
        self.batch = _make_batch(payment_type='one_time', trainer_names='Priya')
        self.student = _make_student('Aarav')
        self.enrollment = BatchEnrollment.objects.create(
            batch=self.batch, student=self.student, joined_date=datetime.date(2026, 7, 1),
        )

    def _get(self, action, user):
        request = self.factory.get(f'/api/batch-enrollments/{self.enrollment.id}/{action}/')
        force_authenticate(request, user=user)
        return BatchEnrollmentViewSet.as_view({'get': action})(request, pk=self.enrollment.id)

    def test_admin_can_download_report(self):
        response = self._get('report', self.admin)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_owning_trainer_can_download_report(self):
        response = self._get('report', self.trainer.user)
        self.assertEqual(response.status_code, 200)

    def test_non_owning_trainer_cannot_download_report(self):
        response = self._get('report', self.other_trainer.user)
        self.assertEqual(response.status_code, 403)

    def test_certificate_blocked_while_batch_is_still_ongoing(self):
        response = self._get('certificate', self.admin)
        self.assertEqual(response.status_code, 400)

    def test_certificate_blocked_for_a_withdrawn_enrollment_even_if_batch_completed(self):
        self.batch.status = 'completed'
        self.batch.save(update_fields=['status'])
        self.enrollment.status = 'withdrawn'
        self.enrollment.save(update_fields=['status'])
        response = self._get('certificate', self.admin)
        self.assertEqual(response.status_code, 400)

    def test_certificate_available_once_batch_completed_and_enrollment_active(self):
        self.batch.status = 'completed'
        self.batch.save(update_fields=['status'])
        response = self._get('certificate', self.admin)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
