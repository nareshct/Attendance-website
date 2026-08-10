import re
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db import transaction
from django.db.models import Count, Sum
from openpyxl import Workbook, load_workbook

from students.models import Student

from .models import PAYMENT_MILESTONES, Batch, BatchEnrollment, BatchInstallment


# Mirrors reports.views.csv_safe() — a cell starting with =, +, -, or @ is executed as
# a formula by Excel when opened. The student name/phone and occupation/email/source
# here originate from a public Google Form (via import_students_from_excel) or the
# enroll form directly, so nothing stops one from starting with one of those characters.
# openpyxl writes these as typed string cells (not CSV plain text), so the practical
# risk is lower than reports/'s CSV exports, but the fix is free and keeps this
# consistent with the rest of the app's exports.
def _xlsx_safe(value):
    text = str(value or '')
    if text and text[0] in ('=', '+', '-', '@'):
        return f"'{text}"
    return text


def create_batch_installments(batch_enrollment):
    """Build a student's BatchInstallment rows for their batch's payment plan.

    The batch has one flat fee (fee_per_student), split evenly across the plan's
    installments — there's no per-class rate here to derive an uneven upfront amount
    from, unlike the individual B2C course plans. due_at_sessions is resolved from the
    plan's milestone fractions x the batch's total_classes at creation time. The last
    installment absorbs any rounding remainder so the installments always sum to
    exactly fee_per_student.
    """
    batch = batch_enrollment.batch
    fractions = PAYMENT_MILESTONES[batch.payment_type]
    n = len(fractions)
    total = batch.fee_per_student
    share = (total / n).quantize(Decimal('0.01'))

    installments = []
    for i, frac in enumerate(fractions, start=1):
        due_at_sessions = None if frac is None else max(1, round(batch.total_classes * frac))
        amount = total - share * (n - 1) if i == n else share
        installments.append(BatchInstallment(
            batch_enrollment=batch_enrollment, sequence=i, due_at_sessions=due_at_sessions, amount=amount,
        ))
    BatchInstallment.objects.bulk_create(installments)
    return installments


def batch_revenue_summary(batch):
    """Revenue for one batch — entirely independent of BillingCycle / Payout /
    ClientInvoice, which only cover individual per-class B2B/B2C billing.

    - gross_expected: the sum of every active enrollment's non-cancelled installments —
      the full amount this batch will bring in once every installment is collected.
      Summed from the actual installment amounts rather than enrolled_count x fee, since
      BatchInstallmentSerializer.validate_amount allows discounting a still-pending
      installment after signup — using the flat fee here would silently drift from what's
      really owed once that happens.
    - collected: every installment actually marked paid, regardless of whether that
      student later withdrew — money genuinely received doesn't stop counting as
      revenue just because they left, unless explicitly refunded (see refunded_amount).
    - pending: still-owed installments for currently *active* enrollments only — a
      withdrawn student's future installments are cancelled, not still expected.
    - net_after_payout: collected minus every payout on this batch (see BatchPayout —
      there can be one per person helping run it), counted whether or not each has
      actually been paid out yet, same accrual convention the rest of the app uses for
      trainer cost — see clients/services.py's our_earning.
    """
    active_enrollments = batch.batch_enrollments.filter(status='active')
    enrolled_count = active_enrollments.count()
    gross_expected = BatchInstallment.objects.filter(
        batch_enrollment__batch=batch, batch_enrollment__status='active',
    ).exclude(paid_status='cancelled').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    paid = BatchInstallment.objects.filter(
        batch_enrollment__batch=batch, paid_status='paid',
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    pending = BatchInstallment.objects.filter(
        batch_enrollment__batch=batch, batch_enrollment__status='active', paid_status='pending',
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_refunded = batch.batch_enrollments.aggregate(total=Sum('refunded_amount'))['total'] or Decimal('0.00')
    collected = paid - total_refunded

    # A cancelled payout was never actually owed, unlike pending/paid which both
    # still count under the accrual convention described above.
    total_payouts = batch.payouts.exclude(paid_status='cancelled').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    return {
        'enrolled_count': enrolled_count,
        'gross_expected': gross_expected,
        'collected': collected,
        'pending': pending,
        'total_refunded': total_refunded,
        'total_payouts': total_payouts,
        'net_after_payout': collected - total_payouts,
    }


def source_report():
    """Enrollment count + revenue broken down by lead_source (how they heard about the
    center) — the ROI signal for ad-driven batch sign-ups. Covers every enrollment, not
    just ones without a linked student — lead_source is lead-tracking info kept
    alongside the (always-linked) student now, not a marker of registration status. An
    enrollment added via the "existing student" picker has no source at all, grouped
    under "Not recorded" here. Two aggregate queries, not one per source."""
    counts = (
        BatchEnrollment.objects.values('lead_source').annotate(enrolled_count=Count('id')).order_by()
    )
    money = (
        BatchInstallment.objects.exclude(paid_status='cancelled')
        .values('batch_enrollment__lead_source', 'paid_status')
        .annotate(total=Sum('amount')).order_by()
    )

    rows = {}

    def bucket(raw_source):
        source = (raw_source or '').strip() or 'Not recorded'
        return rows.setdefault(
            source, {'source': source, 'enrolled_count': 0, 'collected': Decimal('0.00'), 'pending': Decimal('0.00')},
        )

    for row in counts:
        bucket(row['lead_source'])['enrolled_count'] += row['enrolled_count']
    for row in money:
        target = bucket(row['batch_enrollment__lead_source'])
        key = 'collected' if row['paid_status'] == 'paid' else 'pending'
        target[key] += row['total'] or Decimal('0.00')

    return sorted(rows.values(), key=lambda r: r['enrolled_count'], reverse=True)


# Headers this import expects, after the admin renames the columns on a downloaded
# Google Form response sheet to match (extra columns like the form's own Timestamp are
# simply ignored — matching is by header name, not position, and is punctuation/case
# insensitive, so "How did you know about us?" and "how did you know about us" both work).
# "Grade"/"Parent Name"/"Place" are optional — filled in when known, left blank
# otherwise (same as Payment Status) — see find_or_create_student().
IMPORT_COLUMNS = [
    'Name', 'Phone Number', 'Grade', 'Occupation', 'Email', 'Parent Name', 'Place',
    'How did you know about us?', 'Payment Status',
]

PAID_VALUES = {'paid', 'yes', 'y', 'true', '1'}


def _normalize_header(value):
    return re.sub(r'[^a-z0-9]', '', str(value or '').strip().lower())


def _digits_only(value):
    return re.sub(r'\D', '', str(value or ''))


def find_or_create_student(name, phone, grade='', parent_name='', place=''):
    """Resolve a name typed into the batch "add student" form or an Excel import row to
    a real, registered Student — every batch enrollment ends up linked to one, never
    just a name/phone captured on the enrollment itself.

    Matched by case-insensitive name + digit-normalized phone against every registered
    student (not just this batch) — a phone-less name is never matched (too easy to
    false-positive on a common name) and always creates a new Student instead. New
    students are created B2C with no client — a batch isn't tied to one — and whatever
    grade/parent_name/place was given (often blank; batches don't always collect these).
    A match against an existing student never overwrites their existing details with
    whatever was typed this time — those fields are only ever used when creating new.

    Returns (student, created).
    """
    phone_digits = _digits_only(phone)
    if phone_digits:
        for candidate in Student.objects.filter(name__iexact=name.strip()):
            if _digits_only(candidate.parent_phone_number) == phone_digits:
                return candidate, False
    student = Student.objects.create(
        name=name.strip(), grade=grade.strip() if grade else '', parent_phone_number=phone.strip() if phone else '',
        parent_name=parent_name.strip() if parent_name else '', place=place.strip() if place else '',
        source_type='B2C', client=None,
    )
    return student, True


def import_students_from_excel(batch, file_obj):
    """Bulk-enroll students into a batch from an admin-edited Excel export of a Google
    Form response sheet — see IMPORT_COLUMNS for the expected headers.

    Every row is resolved to a real, registered Student via find_or_create_student() —
    matched against everyone already registered, or created fresh — and the enrollment
    is linked to it; the row's occupation/email/source are kept on the enrollment's own
    occupation/contact_email/lead_source fields too, as lead-tracking info alongside the
    real student link (name/phone aren't — Student.name/parent_phone_number are the one
    canonical copy). A duplicate within the same batch (same name + phone resolving to
    the same student) is skipped rather than double-enrolled.

    A "Paid" Payment Status marks only the first (due-at-signup) installment paid —
    matching the real workflow: the parent pays that upfront amount via the form before
    the admin verifies their screenshot and imports the row. Any later installments for
    multi-installment batches stay pending, same as a manually added enrollment — the
    admin marks those from the batch page as the batch progresses.
    """
    if not batch.accepting_enrollments:
        raise ValueError('This batch is closed to new enrollments.')

    try:
        workbook = load_workbook(file_obj, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f'Could not read this as an Excel (.xlsx) file: {exc}')

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise ValueError('The file has no rows.')

        column_index = {}
        for i, header in enumerate(header_row):
            key = _normalize_header(header)
            if key and key not in column_index:
                column_index[key] = i

        if _normalize_header('Name') not in column_index:
            raise ValueError('Missing a "Name" column. Expected headers: ' + ', '.join(IMPORT_COLUMNS))

        def cell(row, header):
            idx = column_index.get(_normalize_header(header))
            if idx is None or idx >= len(row) or row[idx] is None:
                return ''
            return str(row[idx]).strip()

        # Dedup by resolved student identity now, not by raw name/phone text — a row
        # that resolves (via find_or_create_student) to a student already enrolled here,
        # whether from an earlier import row, a manual add, or the registered-student
        # path, is skipped rather than double-enrolled.
        already_enrolled_student_ids = set(
            BatchEnrollment.objects.filter(batch=batch).values_list('student_id', flat=True)
        )

        enrolled = 0
        marked_paid = 0
        skipped = []

        # Without this, an error partway through a large sheet (a bad row, a transient
        # DB error) left however many rows had already been created committed, with no
        # way to tell from the response which rows made it in — a re-run of the same
        # file would then also skip those already-committed rows via
        # already_enrolled_student_ids, making the gap invisible. Wrapping the whole
        # pass in one transaction means a failure rolls back to "nothing imported", so a
        # retry behaves like the first attempt.
        with transaction.atomic():
            for row_num, row in enumerate(rows, start=2):
                if row is None or all(v in (None, '') for v in row):
                    continue

                name = cell(row, 'Name')
                if not name:
                    skipped.append({'row': row_num, 'reason': 'Name is required.'})
                    continue

                phone = cell(row, 'Phone Number')
                grade = cell(row, 'Grade')
                occupation = cell(row, 'Occupation')
                email = cell(row, 'Email')
                parent_name = cell(row, 'Parent Name')
                place = cell(row, 'Place')
                source = cell(row, 'How did you know about us?')
                payment_status = cell(row, 'Payment Status')

                student, _ = find_or_create_student(name, phone, grade, parent_name, place)
                if student.id in already_enrolled_student_ids:
                    skipped.append({'row': row_num, 'reason': f'{name} is already enrolled in this batch.'})
                    continue

                enrollment = BatchEnrollment.objects.create(
                    batch=batch, student=student, contact_email=email,
                    occupation=occupation, lead_source=source, joined_date=date.today(), status='active',
                )
                already_enrolled_student_ids.add(student.id)
                installments = create_batch_installments(enrollment)
                enrolled += 1

                if payment_status.lower() in PAID_VALUES and installments:
                    first = installments[0]
                    first.paid_status = 'paid'
                    first.paid_date = date.today()
                    first.save(update_fields=['paid_status', 'paid_date'])
                    marked_paid += 1

        return {
            'enrolled': enrolled,
            'marked_paid': marked_paid,
            'skipped': skipped,
        }
    finally:
        workbook.close()


def build_import_template():
    """An .xlsx with just the expected header row (plus one example row), for the admin
    to use as a guide when renaming columns on a downloaded Google Form response sheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Batch import'
    sheet.append(IMPORT_COLUMNS)
    sheet.append(['Asha Kumar', '9876543210', '', 'Software engineer', 'asha@example.com', '', '', 'Instagram ad', 'Paid'])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def batch_ids_for_trainer(trainer_name):
    """Batches where `trainer_name` appears in the free-text trainer_names list
    (case-insensitive, comma-split) — trainer_names isn't a relation to the Trainer
    table (see Batch's docstring), so this can't be a normal filter() lookup."""
    name = (trainer_name or '').strip().lower()
    if not name:
        return []
    return [
        b.id for b in Batch.objects.only('id', 'trainer_names')
        if name in [n.strip().lower() for n in b.trainer_names.split(',') if n.strip()]
    ]


def all_batches_summary():
    """Top-line stats across every batch, for the Batches list page header.

    Aggregates directly instead of calling batch_revenue_summary() per batch — looping
    that (4-5 queries each) turns this into an N+1 as the number of batches grows; this
    stays at a fixed handful of queries regardless of batch count.
    """
    batch_count = Batch.objects.count()
    total_students = BatchEnrollment.objects.filter(status='active').count()

    total_paid = BatchInstallment.objects.filter(paid_status='paid').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_pending = BatchInstallment.objects.filter(
        batch_enrollment__status='active', paid_status='pending',
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_refunded = BatchEnrollment.objects.aggregate(total=Sum('refunded_amount'))['total'] or Decimal('0.00')

    return {
        'batch_count': batch_count,
        'total_students': total_students,
        'total_collected': total_paid - total_refunded,
        'total_pending': total_pending,
    }


def build_export_workbook(batch):
    """.xlsx of a batch's current roster — the mirror of the Excel import, for backups,
    WhatsApp broadcast lists, or handing a roster to a trainer."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Enrolled students'
    sheet.append([
        'Name', 'Student ID', 'Phone Number', 'Occupation', 'Email',
        'How did you know about us?', 'Status', 'Joined Date', 'Amount Paid', 'Amount Pending',
    ])
    enrollments = (
        batch.batch_enrollments.select_related('student').prefetch_related('installments').order_by('-joined_date')
    )
    for e in enrollments:
        paid = sum((i.amount for i in e.installments.all() if i.paid_status == 'paid'), Decimal('0.00'))
        pending = sum((i.amount for i in e.installments.all() if i.paid_status == 'pending'), Decimal('0.00'))
        # Occupation/contact_email/lead_source are lead-tracking info kept on the
        # enrollment itself for every row — see find_or_create_student(). Name/phone
        # come straight from the linked Student — the one canonical copy.
        row = [
            _xlsx_safe(e.student.name), e.student.student_id, _xlsx_safe(e.student.parent_phone_number),
            _xlsx_safe(e.occupation), _xlsx_safe(e.contact_email), _xlsx_safe(e.lead_source),
        ]
        row += [e.status, e.joined_date.isoformat(), round(float(paid), 2), round(float(pending), 2)]
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
